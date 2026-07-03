import json
import os
import time
import shutil
import pandas as pd
import warnings
from watchdog.observers import Observer
from watchdog.events import FileSystemEventHandler
from openpyxl import load_workbook
from openpyxl_image_loader import SheetImageLoader

# 🔧 屏蔽无关警告
warnings.filterwarnings("ignore", category=UserWarning, module="openpyxl")

# 增加了 DESC_FULL 映射
COLUMN_MAP = {
    "MARCA": "MARCA",
    "TIPO": "TIPO",
    "PRODUCTO": "PRODUCTO",
    "DESC_FULL": "DESC_FULL",
    "Precio ( USD )": "Precio ( USD )",
    "Imagen_Path": "Imagen_Path",
    "VIENE LA CAJA": "VIENE_LA_CAJA",
    "video_url": "video_url",
    "qcy-yf": "qcy-yf",
    "新到货": "新到货",  # 👈 添加这一行
}

WATCH_FOLDER = os.path.dirname(os.path.abspath(__file__))
SOURCES_FILE = os.path.join(WATCH_FOLDER, "sources.json")

def extract_images_for_excel(xlsx_path: str):
    name = os.path.basename(xlsx_path)
    base = os.path.splitext(name)[0]
    img_folder = f"{base}_images"
    img_folder_path = os.path.join(WATCH_FOLDER, img_folder)
    os.makedirs(img_folder_path, exist_ok=True)

    wb = load_workbook(xlsx_path, data_only=True)
    sheet = wb.active

    img_col_idx = None
    for col in range(1, sheet.max_column + 1):
        v = sheet.cell(row=1, column=col).value
        if v and str(v).strip().lower() == "imagen":
            img_col_idx = col
            break

    if img_col_idx is None:
        return {}

    image_loader = SheetImageLoader(sheet)
    row_to_path = {}
    img_count = 0

    for row in range(2, sheet.max_row + 1):
        cell = sheet.cell(row=row, column=img_col_idx)
        addr = cell.coordinate
        if image_loader.image_in(addr):
            try:
                image = image_loader.get(addr)
                img_count += 1
                filename = f"{base}_{img_count:03d}.png"
                out_path = os.path.join(img_folder_path, filename)
                image.save(out_path)
                rel_path = f"{img_folder}/{filename}"
                row_to_path[row] = rel_path
                print(f"📸 从 {addr} 导出图片 -> {out_path}")
            except:
                pass
    return row_to_path

def excel_to_json(path):
    name = os.path.basename(path)
    if name.startswith("~$") or not name.lower().endswith(".xlsx"):
        return None

    print(f"\n🚀 开始处理表格: {name}")

    try:
        row_to_img = extract_images_for_excel(path)
        df = pd.read_excel(path)

        # --- 新增：清理表头空格，确保匹配 ---
        df.columns = [str(c).strip() for c in df.columns]

        # ✨ 核心改进：拆分【产品名】和【详细介绍】
        titles = []
        descriptions = []
        
        for val in df["PRODUCTO"]:
            raw_text = str(val).strip()
            # 按照换行符拆分内容
            parts = raw_text.split('\n', 1) 
            
            # 第一行是红色标题
            titles.append(parts[0].strip())
            # 剩下的所有内容是详情
            descriptions.append(parts[1].strip() if len(parts) > 1 else "")

        df["PRODUCTO"] = titles
        df["DESC_FULL"] = descriptions

        # 价格兼容处理
        if "Precio ( USD )" not in df.columns:
            for col in df.columns:
                if "precio" in col.lower() and "usd" in col.lower():
                    df["Precio ( USD )"] = df[col]
                    break
        
        if "Precio ( USD )" in df.columns:
            df["Precio ( USD )"] = pd.to_numeric(df["Precio ( USD )"].astype(str).str.replace(r'[^0-9.]', '', regex=True), errors='coerce').fillna(0)

        # --- 新增：针对 VIENE LA CAJA 的模糊匹配逻辑 ---
        if "VIENE LA CAJA" not in df.columns:
            for col in df.columns:
                if "CAJA" in str(col).upper():
                    df["VIENE LA CAJA"] = df[col]
                    break

        def is_invalid(val):
            s = str(val).strip()
            return s.startswith('=') or "DISPIMG" in s or s.lower() == "nan" or s == ""

        if "MARCA" in df.columns:
            df = df[~df["MARCA"].apply(is_invalid)]
        
        imagen_paths = []
        for idx in df.index:
            excel_row_num = idx + 2
            imagen_paths.append(row_to_img.get(excel_row_num, ""))
        df["Imagen_Path"] = imagen_paths

        # 这里的 keys 会包含新的 "VIENE LA CAJA"
        df = df[[c for c in COLUMN_MAP.keys() if c in df.columns]]
        df = df.rename(columns=COLUMN_MAP)

        # ❗ 重要修改点：在转 JSON 前，把所有 NaN 替换为空字符串，解决网页报错
        df = df.fillna("")

        out_name = os.path.splitext(name)[0] + ".json"
        out_path = os.path.join(WATCH_FOLDER, out_name)
        with open(out_path, "w", encoding="utf-8") as f:
            # 使用 pandas 的 to_dict 配合 json.dump 确保格式整洁
            json.dump(df.to_dict(orient="records"), f, ensure_ascii=False, indent=2)
        
        print(f"✅ 处理完成: {out_name} (共生成 {len(df)} 条数据)")
        return out_name
    except Exception as e:
        print(f"❌ 发生错误: {e}")
        return None

def build_sources_manifest():
    entries = []
    products_entry = None
    master_json = os.path.join(WATCH_FOLDER, "data.json")
    if os.path.exists(master_json):
        entries.append({"key": "stock", "label": "STOCK", "desc": "Catálogo de stock actual", "file": "data.json"})

    for name in os.listdir(WATCH_FOLDER):
        if not name.lower().endswith(".json") or name in ("sources.json", "data.json"):
            continue
        base = os.path.splitext(name)[0]
        if name.lower() == "products.json":
            products_entry = {"key": "lista_actual", "label": "LISTA ACTUAL", "desc": "Lista actual", "file": "products.json"}
            continue
        entries.append({"key": base.lower(), "label": base.upper(), "desc": base, "file": name})

    if products_entry: entries.insert(0, products_entry)
    with open(SOURCES_FILE, "w", encoding="utf-8") as f:
        json.dump(entries, f, ensure_ascii=False, indent=2)

class Handler(FileSystemEventHandler):
    def on_modified(self, event):
        if not event.is_directory and event.src_path.endswith(".xlsx"):
            # 延迟 1 秒等待 Excel 释放文件锁
            time.sleep(1)
            if excel_to_json(event.src_path): build_sources_manifest()

if __name__ == "__main__":
    print("=========================================")
    print("🌟 STARPHONE 报表系统启动中...")
    print("=========================================")
    for name in os.listdir(WATCH_FOLDER):
        if name.lower().endswith(".xlsx") and not name.startswith("~$"):
            excel_to_json(os.path.join(WATCH_FOLDER, name))
    build_sources_manifest()
    
    print("\n✨ 数据转换已完成，正在退出并继续同步到 GitHub...")